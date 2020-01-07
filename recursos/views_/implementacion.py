from django.shortcuts import render
from django.http import JsonResponse
from recursos.views_.implementacion_libraries.register import get_register_all, get_register_byId, get_all_coberturas
import pandas as pd
import os
from django.http import HttpResponse
import re
from openpyxl import load_workbook
from openpyxl.styles import Protection, NumberFormatDescriptor
import numpy as np


def getIndicesTable(InputsTarificador, TABLE):
    filter1 = InputsTarificador
    filter1['Variable'] = filter1.index
    filter1 = filter1.reset_index(drop=True)
    filter2 = filter1.loc[filter1['Variable'] == str(TABLE)]['Variable']
    indexMin = filter2.index[0]
    indexMax = filter2.index[-1] + 1
    return indexMin, indexMax


def tan(request):
    username = ''
    email = ''
    first_name = ''
    last_name = ''
    if request.user.is_authenticated:
        username = request.user.username
        email = request.user.email
        first_name = request.user.first_name
        last_name = request.user.last_name
    configurationView = {
        'username': username,
        'email': email,
        'first_name': first_name,
        'last_name': last_name,
        'title': 'Matrices de configuración',
        'area': 'Recursos',
        'herramienta': 'tan-implementacion',
        'file': 'recursos/implementacion/tan.html',
    }
    return render(request, "principal/base.html", configurationView)


def table(request, category):
    log = get_register_all(category)
    data = []
    if len(log) > 0:
        log['DATETIME'] = log['DATETIME'].dt.strftime('%Y-%b-%d <sup style="color: #2196F3;text-shadow: 1px 1px 5px #7372da9e;">%I:%M %p</sup>')
    for i, row in log.iterrows():
        data.append({
            "ID": str(row['ID']),
            "USER_ID": str(row['USER_ID']),
            "FIRST_NAME": str(row['FIRST_NAME']),
            "LAST_NAME": str(row['LAST_NAME']),
            "DATETIME": str(row['DATETIME']),
            "FILE_INPUT": str(row['FILE_INPUT']),
            "FILE_OUTPUT_TAN": str(row['FILE_OUTPUT_TAN']),
            "SUCCESS_TAN": str(row['SUCCESS_TAN']),
            "TOTAL_TIME_TAN": str(row['TOTAL_TIME_TAN']),
            "ERROR_TAN": str(row['ERROR_TAN']),
            "FILE_OUTPUT_BP": str(row['FILE_OUTPUT_BP']),
            "SUCCESS_BP": str(row['SUCCESS_BP']),
            "TOTAL_TIME_BP": str(row['TOTAL_TIME_BP']),
            "ERROR_BP": str(row['ERROR_BP']),
            "STATUS": str(row['STATUS']),
            "TAG": str(row['TAG']),
            "NOMBRE": str(row['NOMBRE']),
            "COLOR": str(row['COLOR'])
        })
    return JsonResponse({'DTable': data}, safe=False)


def formato(request, id):
    # -- Consultar registro
    register = get_register_byId(id)
    register = register.to_records()

    # -- Leer archivo tarificador (input)
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '../static/pricing/tarificadores_inputs/' + register[0][3])
    xlsxFile = pd.ExcelFile(path)
    Input = pd.DataFrame(pd.read_excel(xlsxFile, sheet_name='Input'))
    Input = Input.iloc[1:, :].reset_index(drop=True)
    InputsTarificador = pd.DataFrame(pd.read_excel(xlsxFile, sheet_name='Tarificador'))

    # -- Consultar base de datos de coberturas
    baseCoberturas = get_all_coberturas()

    # -- Consultar productos
    productos = InputsTarificador.loc[InputsTarificador['comision'] == 'Código Producto']
    productos = productos.iloc[:, 1:].reset_index(drop=True).T.reset_index(drop=True)
    productos = productos.dropna()
    # -- Agrupar productos por 'Código Producto'
    productos[1] = productos[0]
    for i, row in productos.iterrows():
        productos[1][i] = row[0][:-1]
    productosAgrupados = productos.groupby([1]).count().reset_index()
    productosAgrupados[0] = productosAgrupados[1]
    for i, row in productosAgrupados.iterrows():
        productosAgrupados[0][i] = row[0] + '1'

    # -- Crear dataframe para configuración de productos -- #
    DataProductos = InputsTarificador.iloc[:, 0:].set_index('comision').T.reset_index(drop=True)

    # -- Inicio filtros para extraer No. indices para la tabla --#
    # -- Valor Asegurado -- #
    indexMin, indexMax = getIndicesTable(InputsTarificador, 'VLR_ASE')
    # -- TIPO COBERTURA -- #
    indexMinTIP_COB, indexMaxTIP_COB = getIndicesTable(InputsTarificador, 'TIP_COB')
    # -- RAMO COD -- #
    indexMinRAM_COD, indexMaxRAM_COD = getIndicesTable(InputsTarificador, 'RAM_COD')
    # -- Reglas de negocio -- #
    indexMinREG_NEG, indexMaxREG_NEG = getIndicesTable(InputsTarificador, 'REG_NEG')
    # -- CUOTAS -- #
    indexMinCUO, indexMaxCUO = getIndicesTable(InputsTarificador, 'CUO')
    # -- Valor a indemnizar; Forma de pago del siniestro -- #
    indexMinVIN_FPS, indexMaxVIN_FPS = getIndicesTable(InputsTarificador, 'VIN_FPS')
    # -- Distribución Prima Neta antes de IVA -- #
    indexMinDIS_IVA, indexMaxDIS_IVA = getIndicesTable(InputsTarificador, 'DIS_IVA')
    # -- BENEFICIARIOS -- #
    indexMinBEN, indexMaxBEN = getIndicesTable(InputsTarificador, 'BEN')
    # -- Periodo de Carencia -- #
    indexMinPER_CAR, indexMaxPER_CAR = getIndicesTable(InputsTarificador, 'PER_CAR')
    # -- Numero de limites de eventos por vigencia -- #
    indexMinLIM_EVE, indexMaxLIM_EVE = getIndicesTable(InputsTarificador, 'LIM_EVE')

    # --------- INICIO Extraer información de todas las coberturas por producto --------- #
    totalCoberturas = pd.DataFrame(columns=['Producto', 'CodProductoTEMP', 'CodCobertura', 'Cobertura', 'FullCobertura', 'TIP_COB', 'RAM_COD', 'REG_NEG',
                                            'TiempoLimite', 'UnidadDeTiempo', 'DiasContLaboral', 'CUO', 'DiasMinimosDeIncapacidad',
                                            'DiasMaximoDeIncapacidad', 'VIN', 'FPS', 'DIS_IVA', 'BEN', 'PER_CAR', 'LIM_EVE', 'i', 'VLR_ASE'])
    contadorCoberturas = 0
    ListadoCodigosProductos = []

    for i, row in productos.iterrows():
        ListadoCodigosProductos.append(str(row[0][:-1]))
        coberturasFiltro1 = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]
        coberturasFiltro1 = coberturasFiltro1.reset_index(drop=True).T.reset_index()
        coberturasFiltro2 = coberturasFiltro1.iloc[indexMin:indexMax].reset_index()
        coberturasFiltro2['IndiceCobertura'] = coberturasFiltro2.index
        coberturasFiltro3 = coberturasFiltro2.loc[coberturasFiltro2[0] > 0]
        # -- Valor Asegurado
        VLR_ASE = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMin:indexMax].reset_index(drop=True)
        # -- Consultar tipo de cobertura
        TipoCobertura = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinTIP_COB:indexMaxTIP_COB].reset_index(drop=True)
        # -- Consultar Cod ramo
        CodRamoCobertura = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinRAM_COD:indexMaxRAM_COD].reset_index(drop=True)
        # -- Consultar Reglas de negocio
        ReglasDeNegocio = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinREG_NEG:indexMaxREG_NEG].reset_index(drop=True)
        # -- Consultar Cuotas
        Cuotas = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinCUO:indexMaxCUO].reset_index(drop=True)
        # -- Consultar Valor a indemnizar; Forma de pago del siniestro
        VIN_FPS = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinVIN_FPS:indexMaxVIN_FPS].reset_index(drop=True)
        # -- Consultar valor de NET %
        DIS_IVA = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinDIS_IVA:indexMaxDIS_IVA].reset_index(drop=True)
        # -- Consular valores de BENEFICIARIOS
        BEN = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinBEN:indexMaxBEN].reset_index(drop=True)
        # -- Consultar valores de Periodo de Carencia
        PER_CAR = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinPER_CAR:indexMaxPER_CAR].reset_index(drop=True)
        # -- Consultar Numero de limites de eventos por vigencia
        LIM_EVE = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMinLIM_EVE:indexMaxLIM_EVE].reset_index(drop=True)

        for j, rowW in coberturasFiltro3.iterrows():
            totalCoberturas.loc[contadorCoberturas] = ''
            totalCoberturas['i'][contadorCoberturas] = i
            totalCoberturas['Producto'][contadorCoberturas] = row[0]
            totalCoberturas['CodProductoTEMP'][contadorCoberturas] = 'Temporal_' + str(i + 1)
            totalCoberturas['Cobertura'][contadorCoberturas] = rowW['comision']
            # -- Consultar ID de cobertura
            CodCobertura = baseCoberturas.loc[baseCoberturas['COBERTURA'] == str(rowW['comision'])].reset_index(drop=True)
            totalCoberturas['CodCobertura'][contadorCoberturas] = CodCobertura['ID'][0]
            totalCoberturas['FullCobertura'][contadorCoberturas] = CodCobertura['FULLCOBERTURA'][0]
            # -- TIP_COB
            totalCoberturas['TIP_COB'][contadorCoberturas] = TipoCobertura[0][rowW['IndiceCobertura']]
            # -- Consultar Cod ramo
            NaneCodRamoCobertura = CodRamoCobertura[0][rowW['IndiceCobertura']]
            CODNaneCodRamoCobertura = NaneCodRamoCobertura.split('.')
            totalCoberturas['RAM_COD'][contadorCoberturas] = CODNaneCodRamoCobertura[0]
            # -- Consultar Reglas de negocio
            totalCoberturas['REG_NEG'][contadorCoberturas] = ReglasDeNegocio[0][rowW['IndiceCobertura']]
            # -- Consultar 'Tiempo límite de Transacción,  Bloqueo,  límite de Hurto o Daño' de cobertura
            totalCoberturas['TiempoLimite'][contadorCoberturas] = CodCobertura['TIEMPO_LIMITE'][0]
            # -- Consultar 'Unidad de Tiempo Item Columna O' de cobertura
            totalCoberturas['UnidadDeTiempo'][contadorCoberturas] = CodCobertura['UNIDAD_TIEMPO'][0]
            # -- Dias de continuidad laboral (Solo se agrega valor fijo 180, siempre y cuando la cobertura sea desempleo)
            totalCoberturas['DiasContLaboral'][contadorCoberturas] = 0
            if str(rowW['comision']) == 'Desempleo':
                totalCoberturas['DiasContLaboral'][contadorCoberturas] = 180
            # -- Consultar Cuotas
            totalCoberturas['CUO'][contadorCoberturas] = Cuotas[0][rowW['IndiceCobertura']]
            # -- Dias Minimos de Incapacidad (Solo aplica para '%Incapacidad Total Temporal%', se coloca 15)
            DiasMinimosDeIncapacidad = 0
            if ("Incapacidad Total Temporal" in str(rowW['comision'])):
                DiasMinimosDeIncapacidad = 15
            totalCoberturas['DiasMinimosDeIncapacidad'][contadorCoberturas] = DiasMinimosDeIncapacidad
            # -- Dias Maximo de Incapacidad (Solo aplica para itt, se coloca 1000)
            DiasMaximoDeIncapacidad = 0
            if ("Incapacidad Total Temporal" in str(rowW['comision'])):
                DiasMaximoDeIncapacidad = 1000
            totalCoberturas['DiasMaximoDeIncapacidad'][contadorCoberturas] = DiasMaximoDeIncapacidad
            # -- Consultar Valor a indemnizar; Forma de pago del siniestro
            BuscarVIN_FPS = VIN_FPS[0][rowW['IndiceCobertura']].split(';')
            totalCoberturas['VIN'][contadorCoberturas] = BuscarVIN_FPS[0]
            totalCoberturas['FPS'][contadorCoberturas] = BuscarVIN_FPS[1]
            # -- Consultar valor de NET %
            totalCoberturas['DIS_IVA'][contadorCoberturas] = DIS_IVA[0][rowW['IndiceCobertura']]
            # -- Consular valores de BENEFICIARIOS
            totalCoberturas['BEN'][contadorCoberturas] = BEN[0][rowW['IndiceCobertura']]
            # -- Consultar valores de Periodo de Carencia
            totalCoberturas['PER_CAR'][contadorCoberturas] = PER_CAR[0][rowW['IndiceCobertura']]
            # -- Consultar Numero de limites de eventos por vigencia
            totalCoberturas['LIM_EVE'][contadorCoberturas] = LIM_EVE[0][rowW['IndiceCobertura']]
            # -- Valor Asegurado
            totalCoberturas['VLR_ASE'][contadorCoberturas] = VLR_ASE[0][rowW['IndiceCobertura']]

            # -- Contador ++
            contadorCoberturas = contadorCoberturas + 1

    # -- Ordenar primero las coberturas Principales
    totalCoberturas = totalCoberturas.sort_values(by=['CodProductoTEMP', 'TIP_COB'], ascending=[True, False])

    # -- Agrupar Coberturas por 'Producto'
    totalCoberturas['ProductoCobNombre'] = totalCoberturas['Producto']
    totalCoberturas['ProductoNombre'] = totalCoberturas['Producto']
    totalCoberturas['Plan'] = totalCoberturas['Producto']
    for i, row in totalCoberturas.iterrows():
        totalCoberturas['ProductoCobNombre'][i] = str(row['CodCobertura']) + row['Producto'][:-1]
        totalCoberturas['ProductoNombre'][i] = row['Producto'][:-1]
        totalCoberturas['Plan'][i] = row['Producto'][-1:]
    totalCoberturasAgrupadas = totalCoberturas
    totalCoberturasAgrupadas = totalCoberturasAgrupadas.drop_duplicates(subset='ProductoCobNombre')
    totalCoberturasAgrupadas = totalCoberturasAgrupadas.reset_index(drop=True)

    # -- Re organizar el id2 (id padre)
    totalCoberturas["i2"] = (totalCoberturas.ProductoNombre != totalCoberturas.ProductoNombre.shift(1)).fillna(0).cumsum()
    totalCoberturas["i2"] = totalCoberturas["i2"] - 1
    totalCoberturas = totalCoberturas.reset_index(drop=True)

    # -- Re organizar el id2 (id padre)
    totalCoberturasAgrupadas["i2"] = (totalCoberturasAgrupadas.Producto != totalCoberturasAgrupadas.Producto.shift(1)).fillna(0).cumsum()
    totalCoberturasAgrupadas["i2"] = totalCoberturasAgrupadas["i2"] - 1

    # --------- FIN EExtraer información de todas las coberturas por producto --------- #

    # ---------------------------------------------------------------------------------------------------------------------------------- #
    # -------------------------------------------------------- Configuración ----------------------------------------------------------- #
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    titlesConfiguracion = [
        'CODIGO REGISTRO',
        'TIPO DE IMPLEMENTACION',
        'CODIGO PRODUCTO',
        'VERSIÓN',
        'SOCIO',
        'COD SOCIO',
        'SIGLA SOCIO',
        'MONEDA',
        'DIAS DE CANCELACION POR MORA',
        'FECHA INICIO PRODUCTO',
        'GRUPO CONFIGURACION',
        'VIGENCIA',
        'PERIODICIDAD',
        'PRODUCTO FINANCIERO/ LINEA INCENTIVOS',
        'CANAL DE VENTA',
        'NOMBRE PRODUCTO',
        'NOMBRE PRODUCTO PIMS',
        'NOMBRE PRODUCTO SOCIO',
        'TIPO PRODUCTO SEGURO',
        'Product Type',
        'ProductCodeMetierType',
        'Product Activity Label Type',
        'Product Family',
        'NO. POLIZA GRUPO CARDIF/ALFA',
        'RANGO DE TOLERANCIA',
        'ENVIO KIT BIENVENIDA',
        'INCENTIVOS',
        'ARMADO CERTIFICADO CARDIF',
        'ARMADO CERTIFICADO SOCIO',
        'PERIODICAS EN PIMS',
        'RENOVACION EN PIMS',
        'APLICA EXCLUYENTES',
        'PRODUCTOS EXCLUYENTES',
        'FORMA DE INCREMENTO IPC',
        'GRUPO DE FACTURACION',
        'TIPO/FORMA DE COBRANZA',
        'FRANQUICIA MEDIO DE PAGO',
        'COD DE CONVENIO',
        'COD DE COMPENSACION',
        'COD TERMINAL',
        'MODELO DE NEGOCIO',
        '% Distribucion primas',
        'Siniestros IPDV',
        'BASE CALCULO PRIMA',
        'ALFA',
        'Cant. Planes',
        'Cant. Coberturas',
        'Cant. Terceros Comisión',
        'Layout Emisión',
        'Layout Novedad',
        'Código Interno PIMS',
        'Código PIMS',
        'ID Job Periódicas',
        'ID Job Renovación',
        'ID Job Cierre Edad',
        'ID Job Cierre Vigencia',
        'GRUPO DE CARGA PRODUCCION COBRA',
        'PERIODICAS EN COBRA'
    ]
    configuracion = pd.DataFrame(columns=titlesConfiguracion)
    for i, row in productosAgrupados.iterrows():
        configuracion.loc[i] = ''
        # -- INICIO Columnas fijas
        configuracion['CODIGO REGISTRO'][i] = '=VLOOKUP(E' + str(i + 3) + ',socio.conf,10,0) & 2019071'
        configuracion['COD SOCIO'][i] = '=VLOOKUP(E' + str(i + 3) + ',socio.conf,3,0)'
        configuracion['SIGLA SOCIO'][i] = '=IFERROR(VLOOKUP(E' + str(i + 3) + ',socio.conf,4,0),"AC")'
        configuracion['MONEDA'][i] = '=VLOOKUP(E' + str(i + 3) + ',socio.conf,6,0)'
        configuracion['DIAS DE CANCELACION POR MORA'][i] = '=VLOOKUP(E' + str(i + 3) + ',socio.conf,7,0)'
        # configuracion['VIGENCIA'][i] = '=HLOOKUP(K' + str(i + 3) + ',grupo.conf,3,0)'
        # configuracion['PERIODICIDAD'][i] = '=HLOOKUP(K' + str(i + 3) + ',grupo.conf,4,0)'
        configuracion['NOMBRE PRODUCTO'][i] = '=C' + str(i + 3) + ' & "_" &  SUBSTITUTE(VLOOKUP(E' + str(i + 3) + ',socio.conf,2,0)," ","_") & "_" &  SUBSTITUTE(T' + str(i + 3) + '," ","_") & "_" & HLOOKUP(K' + str(i + 3) + ',grupo.conf,4,0)  & "_" & VLOOKUP(S' + str(i + 3) + ',productoSeguro.info,2,0) & "_" &  SUBSTITUTE(O' + str(i + 3) + '," ","_")'
        configuracion['NOMBRE PRODUCTO PIMS'][i] = '=C' + str(i + 3) + ' & "_" &  SUBSTITUTE(VLOOKUP(E' + str(i + 3) + ',socio.conf,13,0)," ","_") & "_" &  VLOOKUP(T' + str(i + 3) + ',productType.conf,5,0) & "_" & HLOOKUP(K' + str(i + 3) + ',grupo.conf,4,0) & "_" & VLOOKUP(S' + str(i + 3) + ',productoSeguro.info,2,0) & "_" & SUBSTITUTE(O' + str(i + 3) + '," ","_")'
        configuracion['ProductCodeMetierType'][i] = '=VLOOKUP(T' + str(i + 3) + ',productType.conf,2,0)'
        configuracion['Product Activity Label Type'][i] = '=VLOOKUP(T' + str(i + 3) + ',productType.conf,3,0)'
        configuracion['Product Family'][i] = '=VLOOKUP(T' + str(i + 3) + ',productType.conf,4,0)'
        configuracion['COD DE COMPENSACION'][i] = '=VLOOKUP(E' + str(i + 3) + ',socio.conf,8,0)'
        configuracion['PERIODICAS EN COBRA'][i] = '=IF(AND(HLOOKUP(K' + str(i + 3) + ',grupo.conf,2,0)="Individual",AR' + str(i + 3) + '<>"Valor enviado por el socio (Prima quemada)"),"Si","No")'
        # -- FIN Columnas fijas
        # configuracion['CODIGO PRODUCTO'][i] = 'Temporal_' + str(i + 1)
        # configuracion['CODIGO PRODUCTO'][i] = row[1]
        # -- SOCIO
        Socio = Input.loc[Input['Variable'] == 'Socio']['Valor'].reset_index(drop=True)[0]
        configuracion['SOCIO'][i] = Socio
        # -- VIGENCIA
        Vigencia = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Vigencia'].reset_index(drop=True)[0]
        configuracion['VIGENCIA'][i] = Vigencia
        # -- PERIODICIDAD
        TipoDePagoDePrima = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Periodicidad pago prima'].reset_index(drop=True)[0]
        configuracion['PERIODICIDAD'][i] = TipoDePagoDePrima
        # -- PRODUCTO FINANCIERO/ LINEA INCENTIVOS
        Linea = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Linea'].reset_index(drop=True)[0]
        configuracion['PRODUCTO FINANCIERO/ LINEA INCENTIVOS'][i] = Linea
        # -- CANAL DE VENTA
        CanalDeVenta = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Canal de Venta'].reset_index(drop=True)[0]
        configuracion['CANAL DE VENTA'][i] = CanalDeVenta
        # -- TIPO PRODUCTO SEGURO
        TIPOPRODUCTOSEGURO = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Tipo producto de seguro'].reset_index(drop=True)[0]
        configuracion['TIPO PRODUCTO SEGURO'][i] = TIPOPRODUCTOSEGURO
        # -- Product Type
        ProductType = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Product Type'].reset_index(drop=True)[0]
        configuracion['Product Type'][i] = ProductType
        # -- INCENTIVOS
        """
        PDV = Input.loc[Input['Variable'] == 'PDV']['Valor'].reset_index(drop=True)[0]
        if PDV > 0:
            configuracion['INCENTIVOS'][i] = 'Si'
        else:
            configuracion['INCENTIVOS'][i] = 'No'
        """
        configuracion['INCENTIVOS'][i] = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['INCENTIVOS'].reset_index(drop=True)[0]
        # -- MODELO DE NEGOCIO
        ModeloDeNegocio = Input.loc[Input['Variable'] == 'MODELO DE NEGOCIO']['Valor'].reset_index(drop=True)[0]
        configuracion['MODELO DE NEGOCIO'][i] = ModeloDeNegocio
        # -- % Distribucion primas
        ModeloDeNegocio = Input.loc[Input['Variable'] == '% Distribucion primas']['Valor'].reset_index(drop=True)[0]
        configuracion['% Distribucion primas'][i] = ModeloDeNegocio
        # -- BASE CALCULO PRIMA
        BaseCalculoPrima = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Base Cálculo Prima'].reset_index(drop=True)[0]
        configuracion['BASE CALCULO PRIMA'][i] = BaseCalculoPrima
        # -- Cant. Planes
        CodigoProducto = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Código Producto'].reset_index(drop=True)[0]
        CodigoProducto = CodigoProducto[:-1]
        configuracion['Cant. Planes'][i] = ListadoCodigosProductos.count(CodigoProducto)
        # -- Cant. Coberturas
        Coberturas = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])].reset_index(drop=True).T.reset_index(drop=True).iloc[indexMin:indexMax]
        numeroCoberturas = Coberturas.loc[Coberturas[0] > 0].count()
        configuracion['Cant. Coberturas'][i] = numeroCoberturas[0]
        # -- Cant. Terceros Comisión
        TercerosConComision = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Terceros con comision'].reset_index(drop=True)[0]
        configuracion['Cant. Terceros Comisión'][i] = TercerosConComision
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    # ---------------------------------------------------------- Cobertura ------------------------------------------------------------- #
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    titlesCobertura = [
        'COD REGISTRO',
        'Version',
        'PRODUCTO',
        'COD COBERTURA',
        'COBERTURAS',
        'COBERTURA SUBSCRIBER COD',
        'COBERTURA SUBSCRIBER DESC',
        'PÓLIZA ALFA',
        'GRUPO GIRO ALFA DESC',
        'GRUPO GIRO ALFA COD',
        'Evento PIM\'s',
        'TIPO COBERTURA',
        'RAMO COD',
        'RAMO DESC',
        'RAMO SUBSCRIBER COD',
        'RAMO SUBSCRIBER DESC',
        'COBERTURAS SUNSYSTEM',
        'Reglas de negocio',
        'Grupo relgas de negocio',
        'ASEGURADO',
        'BENEFICIARIOS',
        'Tiempo límite de Transacción,  Bloqueo,  límite de Hurto o Daño',
        'Unidad de Tiempo Item Columna O',
        'Periodo de Carencia',
        'Dias de continuidad laboral',
        'Numero de limites de eventos por vigencia',
        'CUOTAS',
        'PRESCRIPCION DIAS',
        'PRESCRIPCION',
        'Edad minima de ingreso',
        'Edad maxima de ingreso',
        'Edad maxima de permanencia',
        'Dias Minimos de Incapacidad',
        'Dias Maximo de Incapacidad',
        'Permitir pago de siniestros con concurrencia de eventos',
        'Cobertura para Preexistecias',
        'Como establecer valor a indemnizar',
        'Forma de pago del siniestro',
        'Forma de pago del remanente',
        'Número de cuotas a pagar por siniestro',
        'Entidad Bancaria a la cual se paga los siniestros al Banco/Socio',
        'Tipo de Cuenta',
        'Numero de Cuenta',
        'Cierre automático del siniestro después del ultimo pago?',
    ]
    cobertura = pd.DataFrame(columns=titlesCobertura)
    for i, row in totalCoberturasAgrupadas.iterrows():
        cobertura.loc[i] = ''
        # -- INICIO Columnas fijas
        # cobertura['PRODUCTO'][i] = '=+Configuracion!C' + str(row['i'] + 3)
        cobertura['COD REGISTRO'][i] = '=+Configuracion!A' + str(row['i2'] + 3)
        # cobertura['PRODUCTO'][i] = row['ProductoNombre']
        cobertura['COD COBERTURA'][i] = '=VLOOKUP(E' + str(i + 3) + ',cober.info,3,0)'
        cobertura['COBERTURA SUBSCRIBER COD'][i] = '=IFERROR(IF("REASEGURO CESIÓN PARCIAL"<>"REASEGURO TPA","", VLOOKUP(D' + str(i + 3) + ',cober.alfa.info,2,0)),"")'
        cobertura['COBERTURA SUBSCRIBER DESC'][i] = '=IFERROR(IF("REASEGURO CESIÓN PARCIAL"<>"REASEGURO TPA","", VLOOKUP(D' + str(i + 3) + ',cober.alfa.info,3,0)),"")'
        cobertura['GRUPO GIRO ALFA COD'][i] = '=IFERROR(VLOOKUP(I' + str(i + 3) + ',grupoGiroAlfa.info,2,0),"No Aplica")'
        cobertura['Evento PIM\'s'][i] = '=VLOOKUP(E' + str(i + 3) + ',cober.info,5,0)'
        cobertura['RAMO DESC'][i] = '=VLOOKUP(M' + str(i + 3) + ',Ramo.Info,4,0)'
        # cobertura['RAMO SUBSCRIBER COD'][i] = '=IFERROR(IF("REASEGURO CESIÓN PARCIAL"<>"REASEGURO TPA","", VLOOKUP(M' + str(i + 3) + ',Ramo.Info,6,0)),"")'
        # cobertura['RAMO SUBSCRIBER DESC'][i] = '=IFERROR(IF("REASEGURO CESIÓN PARCIAL"<>"REASEGURO TPA","", VLOOKUP(M' + str(i + 3) + ',Ramo.Info,7,0)),"")'
        cobertura['RAMO SUBSCRIBER COD'][i] = '=IF(VLOOKUP(VLOOKUP(C' + str(i + 3) + ',Configuracion!C:E,3,0),tercero.info,14,0)="Reaseguro Perú",VLOOKUP(M' + str(i + 3) + ',Ramo.Info,7,0),IF(VLOOKUP(VLOOKUP(C' + str(i + 3) + ',Configuracion!C:E,3,0),tercero.info,14,0)="Reaseguro AVAL",VLOOKUP(M' + str(i + 3) + ',Ramo.Info,5,0),"No Aplica"))'
        cobertura['RAMO SUBSCRIBER DESC'][i] = '=IF(VLOOKUP(VLOOKUP(C' + str(i + 3) + ',Configuracion!C:E,3,0),tercero.info,14,0)="Reaseguro Perú",VLOOKUP(M' + str(i + 3) + ',Ramo.Info,8,0),IF(VLOOKUP(VLOOKUP(C' + str(i + 3) + ',Configuracion!C:E,3,0),tercero.info,14,0)="Reaseguro AVAL",VLOOKUP(M' + str(i + 3) + ',Ramo.Info,6,0),"No Aplica"))'
        cobertura['COBERTURAS SUNSYSTEM'][i] = '=VLOOKUP(E' + str(i + 3) + ',cober.info,4,0)'
        cobertura['Grupo relgas de negocio'][i] = '=CONCATENATE("FAM",D' + str(i + 3) + ',"-",R' + str(i + 3) + ')'
        cobertura['PRESCRIPCION DIAS'][i] = '=IF(Z' + str(i + 3) + '="EXTRAORDINARIA",1825,IF(Z' + str(i + 3) + '="ORDINARIA",730,0))'
        cobertura['PRESCRIPCION'][i] = '=VLOOKUP(E' + str(i + 3) + ',cober.info,6,0)'
        # -- FIN Columnas fijas
        # cobertura['PRODUCTO'][i] = row['CodProductoTEMP']
        cobertura['COBERTURAS'][i] = row['Cobertura']
        # cobertura['COD COBERTURA'][i] = row['CodCobertura']
        cobertura['TIPO COBERTURA'][i] = row['TIP_COB']
        cobertura['RAMO COD'][i] = int(row['RAM_COD'])
        cobertura['Reglas de negocio'][i] = row['REG_NEG']
        cobertura['ASEGURADO'][i] = 'Persona Titular del producto financiero'
        cobertura['BENEFICIARIOS'][i] = row['BEN']
        cobertura['Tiempo límite de Transacción,  Bloqueo,  límite de Hurto o Daño'][i] = row['TiempoLimite']
        cobertura['Unidad de Tiempo Item Columna O'][i] = row['UnidadDeTiempo']
        cobertura['Periodo de Carencia'][i] = row['PER_CAR']
        cobertura['Dias de continuidad laboral'][i] = row['DiasContLaboral']
        cobertura['Numero de limites de eventos por vigencia'][i] = row['LIM_EVE']
        cobertura['CUOTAS'][i] = row['CUO']
        # -- Edad minima y maxima de ingreso
        EdadMinimaSuscripcion = DataProductos.loc[DataProductos['Código Producto'] == str(row['Producto'])]['Edad mínima suscripción'].reset_index(drop=True)[0]
        EdadMaximaSuscripcion = DataProductos.loc[DataProductos['Código Producto'] == str(row['Producto'])]['Edad máxima suscripción'].reset_index(drop=True)[0]
        EdadMinimaSuscripcion = re.findall('[1-9][1-9] años', EdadMinimaSuscripcion, re.DOTALL)
        EdadMaximaSuscripcion = re.findall('[1-9][1-9] años', EdadMaximaSuscripcion, re.DOTALL)
        cobertura['Edad minima de ingreso'][i] = int(EdadMinimaSuscripcion[0].replace(' años', ''))
        cobertura['Edad maxima de ingreso'][i] = int(EdadMaximaSuscripcion[0].replace(' años', ''))
        # -- Edad maxima de permanencia
        EdadPermanencia = DataProductos.loc[DataProductos['Código Producto'] == str(row['Producto'])]['Edad permanencia'].reset_index(drop=True)[0]
        cobertura['Edad maxima de permanencia'][i] = int(EdadPermanencia.replace(' años + 364 días', ''))
        # -- Dias Minimos de Incapacidad (Solo aplica para '%Incapacidad Total Temporal%', se coloca 15)
        cobertura['Dias Minimos de Incapacidad'][i] = row['DiasMinimosDeIncapacidad']
        # -- Dias Maximo de Incapacidad (Solo aplica para itt, se coloca 1000)
        cobertura['Dias Maximo de Incapacidad'][i] = row['DiasMaximoDeIncapacidad']
        cobertura['Cobertura para Preexistecias'][i] = 'NO'
        cobertura['Como establecer valor a indemnizar'][i] = row['VIN']
        cobertura['Forma de pago del siniestro'][i] = row['FPS']
        cobertura['Forma de pago del remanente'][i] = 'No Aplica'
        cobertura['Número de cuotas a pagar por siniestro'][i] = row['CUO']
        cobertura['Entidad Bancaria a la cual se paga los siniestros al Banco/Socio'][i] = row['BEN']
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    # ----------------------------------------------------------- Tarifa --------------------------------------------------------------- #
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    titlesTarifa = [
        'COD REGISTRO',
        'Version',
        'Producto',
        'Plan',
        'Rango Edad Ini',
        'Rango Edad Fin',
        'Cobetura Codigo',
        'Cobetura Descripción',
        'Ramo',
        'Fecha',
        'Valor Min',
        'Valor Max',
        'Gross %',
        'Taxed %',
        'Net %',
        'Tax %',
        'Gross Cob',
        'Net Cob',
        'Tax Cob',
        'Gross Total',
        'Net Total',
        'Tax Total',
        'PRIMA / TASA - Gross %',
        'PRIMA / TASA - Net %',
        '% IVA Ponderado',
        'Ramo 09 Sustracción',
        'Ramo 24 Desempleo',
        'Ramo 31 Accidentes Personales',
        'Ramo 34 Vida Grupo',
        'Ramo 03 Automóviles',
        'Ramo 25 Hogar',
        'Valores Asegurados Producto - Valor Min',
        'Valores Asegurados Producto - Valor Max'
    ]
    Tarifa = pd.DataFrame(columns=titlesTarifa)
    for i, row in totalCoberturas.iterrows():
        Tarifa.loc[i] = ''

        # -- INICIO Columnas fijas
        # Tarifa['Producto'][i] = '=+Configuracion!C' + str(row['i'] + 3)
        Tarifa['COD REGISTRO'][i] = '=+Configuracion!A' + str(row['i2'] + 3)
        # Tarifa['Producto'][i] = row['ProductoNombre']
        Tarifa['Plan'][i] = row['Plan']
        Tarifa['Gross %'][i] = '=ROUND(N' + str(i + 3) + '/SUMIFS(N:N,C:C,C' + str(i + 3) + ',D:D,D' + str(i + 3) + '),8)'
        Tarifa['Taxed %'][i] = '=O' + str(i + 3) + '*(1+P' + str(i + 3) + ')'
        Tarifa['Tax %'][i] = '=VLOOKUP(I' + str(i + 3) + ',Ramo.Info,5,0)'
        #Tarifa['Gross Cob'][i] = '=ROUND(R' + str(i + 3) + '+S' + str(i + 3) + ',8)'
        Tarifa['Gross Cob'][i] = '=IFERROR(ROUND(T' + str(i + 3) + '*M' + str(i + 3) + ',8),0)'
        #Tarifa['Net Cob'][i] = '=ROUND(U$' + str(i + 3) + '*O' + str(i + 3) + ',8)'
        Tarifa['Net Cob'][i] = '=ROUND(Q' + str(i + 3) + '/(1+P' + str(i + 3) + '),8)'
        #Tarifa['Tax Cob'][i] = '=ROUND(U$' + str(i + 3) + ' * P' + str(i + 3) + ' * O' + str(i + 3) + ',8)'
        Tarifa['Tax Cob'][i] = '=ROUND(Q' + str(i + 3) + '-R' + str(i + 3) + ',8)'
        #Tarifa['Net Total'][i] = '=IF(T' + str(i + 3) + '>0,ROUND(T' + str(i + 3) + '/SUM(N' + str(i + 3) + ':N' + str(i + 3) + '),8),0)'
        Tarifa['Net Total'][i] = '=ROUND(SUM(R' + str(i + 3) + ':R' + str(i + 3) + '),8)'
        #Tarifa['Tax Total'][i] = '=ROUND(T' + str(i + 3) + '-U' + str(i + 3) + ',8)'
        Tarifa['Tax Total'][i] = '=ROUND(SUM(S' + str(i + 3) + ':S' + str(i + 3) + '),8)'
        Tarifa['PRIMA / TASA - Gross %'][i] = '=SUM(M' + str(i + 3) + ':M' + str(i + 3) + ')'
        Tarifa['PRIMA / TASA - Net %'][i] = '=SUM(O' + str(i + 3) + ':O' + str(i + 3) + ')'
        Tarifa['% IVA Ponderado'][i] = '=SUMPRODUCT(O' + str(i + 3) + ':O' + str(i + 3) + ',P' + str(i + 3) + ':P' + str(i + 3) + ')'
        Tarifa['Ramo 09 Sustracción'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"9",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Ramo 24 Desempleo'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"24",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Ramo 31 Accidentes Personales'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"31",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Ramo 34 Vida Grupo'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"34",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Ramo 03 Automóviles'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"3",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Ramo 25 Hogar'][i] = '=SUMIF($I' + str(i + 3) + ':$I' + str(i + 3) + ',"25",$O' + str(i + 3) + ':$O' + str(i + 3) + ')'
        Tarifa['Valores Asegurados Producto - Valor Min'][i] = '=MIN(K' + str(i + 3) + ':K' + str(i + 3) + ')'
        Tarifa['Valores Asegurados Producto - Valor Max'][i] = '=SUM(L' + str(i + 3) + ':L' + str(i + 3) + ')'
        # -- FIN Columnas fijas
        # Tarifa['Producto'][i] = row['CodProductoTEMP']
        Tarifa['Cobetura Codigo'][i] = row['CodCobertura']
        Tarifa['Cobetura Descripción'][i] = row['Cobertura']
        # -- Edad minima y maxima de ingreso
        EdadMinimaSuscripcion = DataProductos.loc[DataProductos['Código Producto'] == str(row['Producto'])]['Edad mínima suscripción'].reset_index(drop=True)[0]
        EdadMaximaSuscripcion = DataProductos.loc[DataProductos['Código Producto'] == str(row['Producto'])]['Edad máxima suscripción'].reset_index(drop=True)[0]
        EdadMinimaSuscripcion = re.findall('[1-9][1-9] años', EdadMinimaSuscripcion, re.DOTALL)
        EdadMaximaSuscripcion = re.findall('[1-9][1-9] años', EdadMaximaSuscripcion, re.DOTALL)
        Tarifa['Rango Edad Ini'][i] = int(EdadMinimaSuscripcion[0].replace(' años', ''))
        Tarifa['Rango Edad Fin'][i] = int(EdadMaximaSuscripcion[0].replace(' años', ''))
        Tarifa['Ramo'][i] = int(row['RAM_COD'])
        # -- Calcular cuota
        if '+' in str(row['CUO']):
            CUOta = row['CUO'].split('+')
            CUOmin = float(CUOta[0])
            CUOmax = float(CUOta[1]) + float(CUOta[0])
            Valor_Min = CUOmin * row['VLR_ASE']
            Valor_Max = CUOmax * row['VLR_ASE']
        else:
            Valor_Min = row['VLR_ASE']
            Valor_Max = row['VLR_ASE']
        Tarifa['Valor Min'][i] = Valor_Min
        Tarifa['Valor Max'][i] = Valor_Max
        Tarifa['Net %'][i] = row['DIS_IVA']
        PrimaCliente = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Prima Cliente'].reset_index(drop=True)
        Tarifa['Gross Total'][i] = PrimaCliente[0]
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    # ---------------------------------------------------------- Comisión -------------------------------------------------------------- #
    # ---------------------------------------------------------------------------------------------------------------------------------- #
    titlesComision = [
        'COD REGISTRO',
        'Version',
        'Producto',
        'Tipo Tercero',
        'Nombre Tercero',
        'NIT',
        'RAZÓN SOCIAL',
        'Tipo Movimiento',
        'Fecha Inicio',
        'Fecha Fin',
        'BASE CALCULO COMISIONES',
        'Comision antes de IVA',
        'IVA',
        'Comision despues de IVA'
    ]
    Comision = pd.DataFrame(columns=titlesComision)
    contador = 0
    for i, row in productosAgrupados.iterrows():
        # -- Calcular números de terceros
        # -- Cant. Terceros Comisión
        TercerosConComision = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Terceros con comision'].reset_index(drop=True)[0]

        for numTer in range(TercerosConComision):
            contador = contador + numTer
            Comision.loc[i + contador] = ''
            # -- INICIO Columnas fijas
            Comision['Producto'][i + contador] = '=+Configuracion!C' + str(i + 3)
            Comision['NIT'][i + contador] = '=VLOOKUP(E' + str(i + contador + 3) + ', tercero.info,11,0)'
            Comision['RAZÓN SOCIAL'][i + contador] = '=VLOOKUP(E' + str(i + contador + 3) + ', tercero.info,12,0)'
            Comision['IVA'][i + contador] = np.where(
                "Alfa" in Input.loc[Input['Variable'] == 'Socio']['Valor'].reset_index(drop=True)[0],
                0,
                1.19
            )
            # -- FIN Columnas fijas
            #Comision['COD REGISTRO'][i] = '=+Configuracion!A' + str(row['i2'] + 3)

            if numTer == 0:  # -- DATOS DEL PRIMER TERCERO
                Comision['Tipo Tercero'][i + contador] = 'Socio'
                Comision['Nombre Tercero'][i + contador] = Input.loc[Input['Variable'] == 'Socio']['Valor'].reset_index(drop=True)[0]
                # -- Comision antes de IVA
                comision = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Comisión Socio'].reset_index(drop=True)[0]
                Comision['Comision antes de IVA'][i + contador] = comision
                Comision['Comision despues de IVA'][i + contador] = Comision['Comision antes de IVA'][i + contador] * (Comision['IVA'][i + contador])
            elif numTer == 1:  # -- DATOS DEL SEGUNDO TERCERO
                Comision['Tipo Tercero'][i + contador] = Input.loc[Input['Variable'] == 'Tipo de Tercero 1']['Valor'].reset_index(drop=True)[0]
                Comision['Nombre Tercero'][i + contador] = Input.loc[Input['Variable'] == 'Intermediario']['Valor'].reset_index(drop=True)[0]
                # -- Comision antes de IVA
                comision = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Comisión Intermediario'].reset_index(drop=True)[0]
                Comision['Comision antes de IVA'][i + contador] = comision / 1.19
                Comision['Comision despues de IVA'][i + contador] = comision
            elif numTer == 2:  # -- DATOS DEL TERCER TERCERO
                Comision['Tipo Tercero'][i + contador] = Input.loc[Input['Variable'] == 'Tipo de Tercero 2']['Valor'].reset_index(drop=True)[0]
                Comision['Nombre Tercero'][i + contador] = Input.loc[Input['Variable'] == 'Facilitador']['Valor'].reset_index(drop=True)[0]
                # -- Comision antes de IVA
                comision = DataProductos.loc[DataProductos['Código Producto'] == str(row[0])]['Comisión Facilitador'].reset_index(drop=True)[0]
                Comision['Comision antes de IVA'][i + contador] = comision / 1.19
                Comision['Comision despues de IVA'][i + contador] = comision

            Comision['BASE CALCULO COMISIONES'][i + contador] = 'Prima Neta'
            # -- IVA
            Comision['IVA'][i + contador] = np.where(
                "Alfa" in Input.loc[Input['Variable'] == 'Socio']['Valor'].reset_index(drop=True)[0],
                0,
                0.19
            )
            # ( Siempre es 19% excepto en alfa )

    # ------------------------------ #
    # ------- Imprimir excel ------- #
    # ------------------------------ #
    # -- Leer archivo base (Plantilla implementación)
    pathBook = os.path.join(os.path.dirname(os.path.dirname(__file__)), '../static/pricing/business_plan_templates/Matriz_de_configuracion3.xlsx')
    book = load_workbook(pathBook)
    # file_output = 'Formato_ID_' + str(id) + '_' + time.strftime("%Y_%m_%d_%H_%M_%S") + '.xlsx'
    file_output = 'Matriz_de_configuración_ID_' + str(id) + '.xlsx'
    path2 = os.path.join(os.path.dirname(os.path.dirname(__file__)), '../static/pricing/formato_implementacion/' + file_output)
    writer = pd.ExcelWriter(path2, engine='openpyxl')
    writer.book.strings_to_formulas = True
    # -- Agregar hojas de archivo base
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # -- Imprimir datos calculados
    configuracion.to_excel(writer, sheet_name='Configuracion', index=None, float_format='%.15f', startrow=2, header=False)
    cobertura.to_excel(writer, sheet_name='Cobertura', index=None, float_format='%.15f', startrow=2, header=False)
    Tarifa.to_excel(writer, sheet_name='Tarifa', index=None, float_format='%.15f', startrow=2, header=False)
    Comision.to_excel(writer, sheet_name='Comision', index=None, float_format='%.15f', startrow=2, header=False)

    protectionStyle = Protection(locked=False)
    # ---------------------------------------------------------------------- #
    # ------- Desbloquear celdas para implementación | Configuración ------- #
    # ---------------------------------------------------------------------- #
    ws = writer.sheets['Configuracion']
    ColumnasDesbloqueadas = {
        'A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'P', 'Q', 'R', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL',
        'AM', 'AN', 'AQ', 'AS', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF'
    }
    for i in ColumnasDesbloqueadas:
        for cell in ws[i]:
            cell.protection = protectionStyle
    # ------------------------------------------------------------------ #
    # ------- Desbloquear celdas para implementación | Cobertura ------- #
    # ------------------------------------------------------------------ #
    ws = writer.sheets['Cobertura']
    ColumnasDesbloqueadas = {
        'A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'N', 'O', 'P', 'Q', 'S', 'AB', 'AC', 'AI', 'AK', 'AL', 'AM', 'AP', 'AQ', 'AR'
    }
    for i in ColumnasDesbloqueadas:
        for cell in ws[i]:
            cell.protection = protectionStyle
    # --------------------------------------------------------------- #
    # ------- Desbloquear celdas para implementación | Tarifa ------- #
    # --------------------------------------------------------------- #
    ws = writer.sheets['Tarifa']
    ColumnasDesbloqueadas = {
        'A', 'B', 'C', 'D', 'J', 'M', 'N', 'P', 'Q', 'R', 'S', 'U', 'V', 'W', 'X', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG'
    }
    for i in ColumnasDesbloqueadas:
        for cell in ws[i]:
            cell.protection = protectionStyle
    # ----------------------------------------------------------------- #
    # ------- Desbloquear celdas para implementación | Comisión ------- #
    # ----------------------------------------------------------------- #
    ws = writer.sheets['Comision']
    ColumnasDesbloqueadas = {
        'A', 'B', 'C', 'F', 'G', 'H', 'I', 'J'
    }
    for i in ColumnasDesbloqueadas:
        for cell in ws[i]:
            cell.protection = protectionStyle

    # -- Formato para valores en hoja tarifa
    ws = writer.sheets['Tarifa']
    ColumnasDesbloqueadas = {
        'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG'
    }
    for i in ColumnasDesbloqueadas:
        for cell in ws[i]:
            cell.number_format = '0.00000000'

    writer.save()
    # -- Imprimir formato
    with open(path2, 'rb') as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename(path2)
        return response
