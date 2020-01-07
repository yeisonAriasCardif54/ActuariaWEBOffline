import pandas as pd
import time
import pandas.io.formats.excel
import os
from openpyxl import load_workbook
import numpy as np
import comtypes.client as cl
import win32com.client as win32
import win32com

from pricing.views_.calcular_bp.libraries.obtener_archivo import generateExcel
from pricing.views_.calcular_bp.libraries.obtener_nuevas_columnas import obtener_nuevas_columnas_input
from pricing.views_.calcular_bp.libraries.obtener_nuevas_columnas import obtener_nuevas_columnas_output
from pricing.views_.calcular_bp.libraries.obtener_costos_adquisicion import obtener_costos_unitarios_adquisicion
from pricing.views_.calcular_bp.libraries.obtener_costos_adquisicion import obtener_costos_adquisicion
from pricing.views_.calcular_bp.libraries.obtener_bp_mensualYanual import obtener_bp_mensual
from pricing.views_.calcular_bp.libraries.obtener_bp_mensualYanual import obtener_bp_anual
from pricing.views_.calcular_bp.libraries.obtener_nuevos import obtener_nuevos
from pricing.views_.calcular_bp.libraries.obtener_tasa_caida import obtener_tasa
from pricing.views_.calcular_bp.libraries.obtener_vigentes_cancelaciones import obtener_vigentes_y_cancelaciones
from pricing.views_.calcular_bp.libraries.obtener_siniestros import obtener_siniestros
from pricing.views_.calcular_bp.libraries.obtener_vlrprima_c_d import obtener_vlrprima_c_d
from pricing.views_.calcular_bp.libraries.obtener_primas_emitidas import obtener_primas_emitidas
from pricing.views_.calcular_bp.libraries.obtener_comisiones import obtener_comisiones
from pricing.views_.calcular_bp.libraries.solvencyii.getNonLifeSCR import NonLifeSCR
from pricing.views_.calcular_bp.libraries.solvencyii.calcular_CalcularInputs import SolvencyII_CalcularInputs
from pricing.views_.calcular_bp.libraries.solvencyii.calcular_Data import solvencyII_Data
from pricing.views_.calcular_bp.libraries.solvencyii.calcular_mk_shocks import solvencyII_MK_Shocks
from pricing.views_.calcular_bp.libraries.solvencyii.obtener_archivo import generateExcelSolvencyII
from pricing.views_.calcular_bp.libraries.solvencyii.get_shock_parameters import getShockParameters

pd.options.display.float_format = '{:,.15f}'.format

pandas.io.formats.excel.header_style = None


def get_data_bp(file, ver_mas_detalles):
    meses = 120
    start_time_GLOBAL = time.time()
    # -- Lectura de inputs -- #
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '../../' + file)
    xlsxFile = pd.ExcelFile(path)
    Inputs = pd.DataFrame(pd.read_excel(xlsxFile, sheet_name='Inputs'))

    # -- Calcular nuevas columnas -- #
    Inputs = obtener_nuevas_columnas_input(Inputs)
    # -- Extraer información de costo de adquisición -- #
    Inputs = obtener_costos_unitarios_adquisicion(Inputs)
    # -- Generar nuevos -- #
    OutPut = obtener_nuevos(Inputs)
    # -- Obtener tasa de caída -- #
    OutPut, OutPut_tasa = obtener_tasa(OutPut, meses, 1)
    # -- Obtener vigentes y cancelaciones -- #
    OutPut, OutPut_tasa_caida_cancel = obtener_vigentes_y_cancelaciones(OutPut, OutPut_tasa, meses, 0)
    # -- Obtener Siniestros -- #
    OutPut = obtener_siniestros(OutPut)
    # -- Obtener vlrprimac y vlrprimad para posterior cálculos de primas -- #
    OutPut, OutPut_vlrprimac, OutPut_vlrprimad = obtener_vlrprima_c_d(OutPut, meses)
    # -- Obtener primas emitidas (GWP) -- #
    OutPut, OutPut_PRI_temp = obtener_primas_emitidas(OutPut, meses, OutPut_tasa_caida_cancel, OutPut_tasa, OutPut_vlrprimac, OutPut_vlrprimad)
    # -- Obtener comisiones -- #
    OutPut = obtener_comisiones(OutPut)
    # -- Nuevas columnas en OutPut -- #
    OutPut = obtener_nuevas_columnas_output(OutPut)
    # -- Nuevas columnas en OutPut de costos de adquisición -- #
    OutPut = obtener_costos_adquisicion(OutPut)
    # -- Generar OutPut detalles -- #

    OutPut_imprimir = pd.DataFrame(
        {
            'Producto': OutPut['Producto'],
            'Tipo de prima': OutPut['Tipo de prima'],
            'Canal': OutPut['Canal'],
            'Stock inicial': OutPut['Stock inicial'],
            'Clientes potenciales/mes': OutPut['Clientes potenciales/mes'],
            'Penetración': OutPut['Penetración'],
            'Tasa crecimiento mensual': OutPut['Tasa crecimiento mensual'],
            'Se alcanza la penetración en No. Meses': OutPut['Se alcanza la penetración en No. Meses'],
            'Duración de producción en meses': OutPut['Duración de producción en meses'],
            'Caída': OutPut['Caida'],
            'Duración del producto financiero': OutPut['Duración del producto financiero'],
            'fecha': OutPut['fecha'],
            'Mes': OutPut['Mes'],
            'Vlr prima': OutPut['Vlr prima'],
            'porcentaje_penetración': OutPut['porcentaje_penetración'],
            'penetración': OutPut['penetración'],
            'nuevos': OutPut['nuevos'],
            'vigentes': OutPut['vigentes'],
            'cancelaciones': OutPut['cancelaciones'],
            'siniestros': OutPut['siniestros'],
            'GWP': OutPut['gwp'],
            'GWP Cancelaciones': OutPut['gwpn'],
            'UPR_eop': OutPut['upr'],
            'Earned_Premium': OutPut['earnedP'],
            'DAC': OutPut['dac'],
            'Total Earned_Commissions': OutPut['earnedC'],
            'Partner Earned_Commissions': OutPut['ecs'],
            'Broker Earned_Commissions': OutPut['ecb'],
            'Overheads': OutPut['Overheads'],
            '- change in UPR net of DAC': OutPut['- change in UPR net of DAC'],
            '- Premium Refund': OutPut['- Premium Refund'],
            '+ Technical Interest On Unearned Premium Reserve': OutPut['+ Technical Interest On Unearned Premium Reserve'],
            '- Insurer Written Acquisition Costs Loading': OutPut['- Insurer Written Acquisition Costs Loading'],
            '- Insurer Earned Acquisition Costs Loading': OutPut['- Insurer Earned Acquisition Costs Loading'],
            '- Insurer Earned Operating Expenses Loading': OutPut['- Insurer Earned Operating Expenses Loading'],
            '- Earned Commission': OutPut['- Earned Commission'],
            '- Earned Insurer Capital Cost Loading': OutPut['- Earned Insurer Capital Cost Loading'],
            '- Insurer Unearned Acquisition Costs Loading': OutPut['- Insurer Unearned Acquisition Costs Loading'],
            # '- Insurer Writing Acquisition Costs Loading': OutPut['- Insurer Writing Acquisition Costs Loading'],
            'Total Earned Risk Premium': OutPut['Total Earned Risk Premium'],
            'Incurred Claims': OutPut['Incurred Claims'],
            'Paid Claims': OutPut['Paid Claims'],
            '+ Change in Claim Reserve': OutPut['+ Change in Claim Reserve'],
            'Claim Reserves': OutPut['Claim Reserves'],
            'Technical Result': OutPut['Technical Result'],
            'Pure Technical Loss Ratio': OutPut['Pure Technical Loss Ratio'],
            'Technical Result with CoC': OutPut['Technical Result with CoC'],

            'Technical Result with CoC Xpu': OutPut['Technical Result with CoC Xpu'],
            'Technical Result with CoC sumXpu': OutPut['Technical Result with CoC sumXpu'],
            'Technical Result with CoC sumXpu ANIO': OutPut['Technical Result with CoC sumXpu ANIO'],
            'Technical Result with CoC sumXpu ANIO POS': OutPut['Technical Result with CoC sumXpu ANIO POS'],

            'Variable Commission': OutPut['Variable Commission'],

            'business line': OutPut['business line'],
            'group of partners': OutPut['group of partners'],

            'unit_costs_Acquisition Fixed': OutPut['unit_costs_Acquisition Fixed'],
            'unit_costs_Acquisition Variable': OutPut['unit_costs_Acquisition Variable'],
            'unit_costs_Claims Variable': OutPut['unit_costs_Claims Variable'],
            'unit_costs_Administration Fixed - Direct': OutPut['unit_costs_Administration Fixed - Direct'],
            'unit_costs_Administration Fixed - Structure': OutPut['unit_costs_Administration Fixed - Structure'],
            'unit_costs_Administration Variable': OutPut['unit_costs_Administration Variable'],
            'unit_costs_FTS FTG': OutPut['unit_costs_FTS FTG'],
            'unit_costs_Incidence rate': OutPut['unit_costs_Incidence rate'],

            'Inflation': OutPut['Inflation'],
            '# Agreements': OutPut['# Agreements'],

            '# In force policies': OutPut['# In force policies'],
            'Acquisition - Fixed': OutPut['Acquisition - Fixed'],
            'Acquisition - Variable': OutPut['Acquisition - Variable'],
            'Claims': OutPut['Claims'],
            'Administration Fixed - Direct': OutPut['Administration Fixed - Direct'],
            'Administration Variable': OutPut['Administration Variable'],
            'Administration Fixed - Structure': OutPut['Administration Fixed - Structure'],
            'FTS FTG': OutPut['FTS FTG'],
            'Calc Inflation': OutPut['Calc Inflation'],

            'gwpsc': OutPut['gwpsc'],
            'Impuestos ICA': OutPut['Impuestos ICA'],
            'Impuestos GMF': OutPut['Impuestos GMF'],
            'TOTAL ICA GMF': OutPut['TOTAL ICA GMF'],
            'TOTAL RECAUDO': OutPut['TOTAL RECAUDO'],
            'ASISTENCIA': OutPut['ASISTENCIA'],

            'Costo Capacitación mes': OutPut['Costo Capacitación mes'],
            'Publicidad mes': OutPut['Publicidad mes'],
            'Bolsa Premios mes': OutPut['Bolsa Premios mes'],
            'Costos Marketing': OutPut['Costos Marketing'],

            'Gestores': OutPut['Gestores'],
            'Duración del seguro': OutPut['Duración del seguro'],
            'CalculoPeriodicidad': OutPut['CalculoPeriodicidad'],
            'Costo de Adquisicion Real': OutPut['Costo de Adquisicion Real'],
            'Acquisition cost': OutPut['Acquisition cost'],

            'Investment Rate anual': OutPut['Investment Rate anual'],
            'Investment Rate mensual': OutPut['Investment Rate mensual'],
            'Financial Income on Reserves': OutPut['Financial Income on Reserves'],

            'RT - OVERHEADS': OutPut['RT - OVERHEADS'],
            'NBI': OutPut['NBI'],
            'OVERHEADS': OutPut['OVERHEADS'],
            'GOI': OutPut['GOI'],
            'TAX': OutPut['TAX'],
            'NOI': OutPut['NOI'],

            '% Life': OutPut['% Life'],
            '% Non Life': OutPut['% Non Life'],
            'GWP net of IVA': OutPut['GWP net of IVA'],
            'Premium Life': OutPut['Premium Life'],
            'Premium Non Life': OutPut['Premium Non Life'],
            'Claim Life': OutPut['Claim Life'],
            'Claim non life': OutPut['Claim non life'],
            'Technical Reserves Life': OutPut['Technical Reserves Life'],

            '# in-force insured life': OutPut['# in-force insured life'],
            'Capital (KCOP)': OutPut['Capital (KCOP)'],
            'Sum at risk Life': OutPut['Sum at risk Life'],
            'Solvency margin life': OutPut['Solvency margin life'],
            '% Life Premium': OutPut['% Life Premium'],

            'Component Claims': OutPut['Component Claims'],
            'Component Premium': OutPut['Component Premium'],
            'Solvency margin non life': OutPut['Solvency margin non life'],

            'EquitySUM': OutPut['EquitySUM'],
            'Equity': OutPut['Equity'],
            'Equity/Premium': OutPut['Equity/Premium'],
            'Avg equity': OutPut['Avg equity'],
            'PV Avg equity': OutPut['PV Avg equity'],

            'On Equity (Participate only in IROE calculations)': OutPut['On Equity (Participate only in IROE calculations)'],
            'GOI con PF sobre equity': OutPut['GOI con PF sobre equity'],
            'TAX con PF sobre equity': OutPut['TAX con PF sobre equity'],
            'NOI con PF sobre equity': OutPut['NOI con PF sobre equity'],
            'NOI con PF sobre equitySUM': OutPut['NOI con PF sobre equitySUM'],
            'PV NOI con PF sobre equity': OutPut['PV NOI con PF sobre equity'],

            'TMPYear': OutPut['Year'] / 12,
            'Discount Rate mensual': OutPut['Discount Rate mensual'],
            'Discount Rate annual': OutPut['Discount Rate annual'],
            'Discount Rate': OutPut['Discount Rate'],
            'IROE': OutPut['IROE'],
            'Cash-flow': OutPut['Cash-flow'],
            'IRR': OutPut['IRR'],
            'PVFP': OutPut['PVFP'],
            'Efficiency Ratio': OutPut['Efficiency Ratio'],

            '(RC_(i-1)-RC_i)': OutPut['(RC_(i-1)-RC_i)'],
            'NOI EQ + (RC_(i-1)-RC_i)': OutPut['NOI EQ + (RC_(i-1)-RC_i)'],
            'Value Creation': OutPut['Value Creation'],

            'Exchange rate': OutPut['Exchange rate'],

            'Paid insurer underwriting': OutPut['Paid insurer underwriting'],
            'Fixed Costs': OutPut['Fixed Costs'],
            'Acquisition Costs': OutPut['Acquisition Costs'],
            'Gross Operating Income (Including Financial Margin)': OutPut['Gross Operating Income (Including Financial Margin)'],
            'Net Operating Income (Including Financial Margin)': OutPut['Net Operating Income (Including Financial Margin)'],
            'Technical Loss Ratio': OutPut['Technical Loss Ratio'],

            'Written Commission Income': OutPut['Written Commission Income'],

            'Premium Collection Credit Risk': OutPut['Premium Collection Credit Risk'],
            'Claims Cash Advance Credit Risk': OutPut['Claims Cash Advance Credit Risk'],
            'Reinsurance Credit Risk': OutPut['Reinsurance Credit Risk'],
            'Surrender Credit Risk': OutPut['Surrender Credit Risk'],
            'Total Credit Risk': OutPut['Total Credit Risk'],
        }
    )

    # -- Generar Business Plan Mensual -- #
    OutPut_productsBP_mensual = obtener_bp_mensual(OutPut_imprimir, Inputs)
    # -- Generar Business Plan Anual -- #
    OutPut_productsBP_anual = obtener_bp_anual(OutPut_productsBP_mensual)
    # -- Guardar archivo -- #
    filename_output = 'Business_plan_' + time.strftime("%Y_%m_%d_%H_%M_%S") + '.xlsx'
    path2 = os.path.join(os.path.dirname(os.path.dirname(__file__)), '../../static/pricing/business_plan_output/' + filename_output)
    writer = pd.ExcelWriter(path2, engine='xlsxwriter')
    writer = generateExcel(writer, ver_mas_detalles, OutPut_imprimir, OutPut_tasa, OutPut_tasa_caida_cancel, OutPut_productsBP_mensual, OutPut_productsBP_anual, OutPut_vlrprimac, OutPut_vlrprimad, OutPut_PRI_temp, Inputs)

    # ------------------------------ #
    # -- Cálculos de solvencia II -- #
    # ------------------------------ #
    # -- MK Shocks
    # MK_Shocks = solvencyII_MK_Shocks()
    # -- Inputs Solvencia
    InputsSolvencia = SolvencyII_CalcularInputs(xlsxFile, Inputs, OutPut, OutPut_productsBP_anual)

    # -- Data
    # -- DataSolvencia = solvencyII_Data(InputsSolvencia, OutPut)
    # -- Get Shock parameters
    # shockParameters = getShockParameters()
    # -- NonLife SCR -- #
    # InputsNonLifeSCR = NonLifeSCR(InputsSolvencia, shockParameters, OutPut_productsBP_anual, MK_Shocks)
    # writer = generateExcelSolvencyII(writer, InputsSolvencia, shockParameters, InputsNonLifeSCR, MK_Shocks)

    # ----------------------------------------------------------------------------------- #
    # -- INICIO nuevos Cálculos de solvencia II (Version beta de lectura de plantilla) -- #
    # ----------------------------------------------------------------------------------- #

    # -- Leer archivo base (Plantilla implementación)
    folder = 'static/pricing/business_plan_templates'
    pathBook = 'Protection_Global Exposed Amount Calculation Tool (Solvency II) v2 ORIGINAL.xlsm'
    file = folder + '/' + pathBook
    book = load_workbook(file, keep_vba=True)
    # book = load_workbook(pathBook)
    file_output = 'static/pricing/solvenciaII/' + 'TemporalSII' + time.strftime("%Y_%m_%d_%H_%M_%S") + '.xlsm'
    writer2 = pd.ExcelWriter(file_output, engine='openpyxl')
    # -- Agregar hojas de archivo base
    writer2.book = book
    writer2.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir GlobalTable
    # ----------------------------------------------------------------------------------- #
    GlobalTable = InputsSolvencia['GlobalTable'].T[1]
    GlobalTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=2, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir ProductTable
    # ----------------------------------------------------------------------------------- #
    ProductTable = InputsSolvencia['ProductTable'].T[1]
    ProductTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=2, startcol=4, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Asset by rating
    # ----------------------------------------------------------------------------------- #
    AssetByRatingTable = InputsSolvencia['AssetByRatingTable'].T.filter([1, 2])
    AssetByRatingTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=19, startcol=4, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Risk Premium repartition
    # ----------------------------------------------------------------------------------- #
    RiskPremiumRepartition1TableT1 = InputsSolvencia['RiskPremiumRepartition1Table'].filter(['Life repartition: Death']).T[1]
    RiskPremiumRepartition1TableT2 = InputsSolvencia['RiskPremiumRepartition1Table'].filter(['Health repartition (SLT & Non SLT): TD',
                                                                                             'Health repartition (SLT & Non SLT): TPD',
                                                                                             'Health repartition (SLT & Non SLT): Accidental  Death',
                                                                                             'Health repartition (SLT & Non SLT): Hospitalization']).T[1]
    RiskPremiumRepartition1TableT3 = InputsSolvencia['RiskPremiumRepartition1Table'].filter(['"Motor"', '"Fire & Other damage"', '"Misc." Subject to Cat']).T[1]
    RiskPremiumRepartition1TableT1.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=3, startcol=7, header=False)
    RiskPremiumRepartition1TableT2.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=6, startcol=7, header=False)
    RiskPremiumRepartition1TableT3.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=12, startcol=7, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Exposed Amount
    # ----------------------------------------------------------------------------------- #
    ExposedAmountTableT1 = InputsSolvencia['ExposedAmountTable'].filter(['Initial stock of policies', 'Monthly number of new policies on a cruse-speed basis']).T[1]
    ExposedAmountTableT2 = InputsSolvencia['ExposedAmountTable'].filter(['Confidence in technical basis quality', 'Confidence in information quality', 'Upfront Payment']).T[1]
    ExposedAmountTableT1.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=19, startcol=10, header=False)
    ExposedAmountTableT2.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=23, startcol=10, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Loss Ratios (PTLR)
    # ----------------------------------------------------------------------------------- #
    LossRatiosPTLRTable = InputsSolvencia['LossRatiosPTLRTable'].T[1]
    LossRatiosPTLRTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=3, startcol=10, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Claim management overheads
    # ----------------------------------------------------------------------------------- #
    ClaimManagementOverheadsTable = InputsSolvencia['ClaimManagementOverheadsTable'].T[1]
    ClaimManagementOverheadsTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=10, startcol=10, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Exclusion Period
    # ----------------------------------------------------------------------------------- #
    ExclusionPeriodTable = InputsSolvencia['ExclusionPeriodTable'].T[1]
    ExclusionPeriodTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=2, startcol=13, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir TopKapiTable
    # ----------------------------------------------------------------------------------- #
    TopKapiTable = InputsSolvencia['TopKapiTable'].T[1]
    TopKapiTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=6, startcol=14, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir Production & Product
    # ----------------------------------------------------------------------------------- #
    ProductionYProductTable = InputsSolvencia['ProductionYProductTable'].T
    del (ProductionYProductTable[0])
    ProductionYProductTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=29, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Imprimir PremiumDataTable
    # ----------------------------------------------------------------------------------- #
    PremiumDataTable = InputsSolvencia['PremiumDataTable'].T
    del (PremiumDataTable[0])
    PremiumDataTable.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=34, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- bop Premium Reserve net of DAC
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['- change in UPR net of DAC'])]
    del (OutPut_productsBP_anualFILTER['Producto'])
    del (OutPut_productsBP_anualFILTER['item'])
    OutPut_productsBP_anualFILTER.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=39, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Earned Premium
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER2 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['Total Earned Risk Premium'])]
    del (OutPut_productsBP_anualFILTER2['Producto'])
    del (OutPut_productsBP_anualFILTER2['item'])
    OutPut_productsBP_anualFILTER2.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=46, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Earned Insurer's loadings (Opex + Margin)
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER3 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['- Earned Insurer Capital Cost Loading'])]
    del (OutPut_productsBP_anualFILTER3['Producto'])
    del (OutPut_productsBP_anualFILTER3['item'])
    OutPut_productsBP_anualFILTER3.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=47, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Total Claim Incurred
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER4 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['Incurred Claims'])]
    del (OutPut_productsBP_anualFILTER4['Producto'])
    del (OutPut_productsBP_anualFILTER4['item'])
    OutPut_productsBP_anualFILTER4.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=52, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Overheads (Real costs)
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER5 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['OVERHEADS'])]
    del (OutPut_productsBP_anualFILTER5['Producto'])
    del (OutPut_productsBP_anualFILTER5['item'])
    OutPut_productsBP_anualFILTER5.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=59, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Gross Operating Income (including capital remuneration)
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER6 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['GOI'])]
    del (OutPut_productsBP_anualFILTER6['Producto'])
    del (OutPut_productsBP_anualFILTER6['item'])
    OutPut_productsBP_anualFILTER6.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=61, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- eop Solvency I Required Capital
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER8 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['Avg equity'])]
    del (OutPut_productsBP_anualFILTER8['Producto'])
    del (OutPut_productsBP_anualFILTER8['item'])
    OutPut_productsBP_anualFILTER8.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=64, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- bop Solvency I Required Capital
    # ----------------------------------------------------------------------------------- #
    OutPut_productsBP_anualFILTER9 = OutPut_productsBP_anual.loc[OutPut_productsBP_anual['item'].isin(['Avg equity'])]
    del (OutPut_productsBP_anualFILTER9['Producto'])
    OutPut_productsBP_anualFILTER9['item'] = 0
    OutPut_productsBP_anualFILTER9.to_excel(writer2, sheet_name='Inputs', index=None, float_format='%.15f', startrow=65, startcol=1, header=False)

    # ----------------------------------------------------------------------------------- #
    # -- Guardar
    # ----------------------------------------------------------------------------------- #
    writer2.save()

    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filename = os.path.join(fileDir, '../ActuariaOffline/' + file_output)
    filename = os.path.abspath(os.path.realpath(filename))

    """
    progId = "Excel.Application"
    xl = cl.CreateObject(progId)
    wb = xl.Workbooks.Open(r"" + filename)
    wb.Save()
    xl.DisplayAlerts = True
    xl.Quit()
    """
    """
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Application.Run("FilterLoans")
    workbook = excel.Workbooks.Open(r"" + filename)
    workbook.Save()
    workbook.Close()
    excel.Quit()
    """
    """
    import ift.xlwings as xw

    # Start Excel app (invisibly in the background)
    app = xw.App(visible=False)

    # Load excel file into active Excel app
    book = app.books.open(r"Book1.xlsm")

    # Instruct Excel to execute the pre-existing excel macro named "CleanUpMacro"
    book.macro("CleanUpMacro")()

    # Instruct Excel to write a cell value in the first sheet
    book.sheets["Sheet1"].range('A1').value = 42

    # Save workbook and terminate Excel application
    book.save()
    book.close()
    app.kill()
     """

    Outputs_SII = pd.DataFrame(pd.read_excel(filename, sheet_name='Outputs'))

    # -------------------------------------------------------------------------------- #
    # -- FIN nuevos Cálculos de solvencia II (Version beta de lectura de plantilla) -- #
    # -------------------------------------------------------------------------------- #

    writer = generateExcelSolvencyII(writer, InputsSolvencia, Outputs_SII)
    # ---------------------------------- #
    # -- FIN Cálculos de solvencia II -- #
    # ---------------------------------- #

    writer.save()
    # -- Retornar respuesta -- #
    total_time = (time.time() - start_time_GLOBAL)
    total_time = format(total_time, '.2f')
    # return 'Se genero con éxito el archivo en ' + str(total_time) + ' segundos.', path2
    return str(total_time), filename_output
